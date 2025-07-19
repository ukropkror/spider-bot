import asyncio
import random
import pandas as pd
import openpyxl 
from telethon import TelegramClient, events
from datetime import datetime
from collections import deque
import threading
import re
import json
import os
from telethon.errors import UsernameInvalidError, UsernameNotOccupiedError, UserPrivacyRestrictedError, FloodWaitError, PeerFloodError


def is_qualified(user_id: str) -> bool:
    user_id = str(user_id)

    # Проверяем users_database.xlsx
    try:
        df_users = pd.read_excel("users_database.xlsx", header=None)
        for i in range(len(df_users)):
            if str(df_users.iloc[i, 0]) == user_id:
                # Уже есть — ставим неквал
                wb = openpyxl.load_workbook("users_database.xlsx")
                ws = wb.active
                col_count = ws.max_column
                if col_count < 6:
                    ws.cell(row=i+1, column=6).value = "Нет"
                wb.save("users_database.xlsx")
                return False
    except Exception as e:
        print(f"⚠️ Не удалось прочитать users_database.xlsx: {e}")

    # Проверяем leads.xlsx
    try:
        df_leads = pd.read_excel("leads.xlsx")
        match = df_leads[df_leads["ID"].astype(str) == user_id]
        if not match.empty:
            index = match.index[0]
            if match.iloc[0]["Отправлено"] is True:
                # Уже отправлено
                wb = openpyxl.load_workbook("leads.xlsx")
                ws = wb.active
                if "Квалифицирован" not in df_leads.columns:
                    ws.cell(row=1, column=len(df_leads.columns)+1).value = "Квалифицирован"
                ws.cell(row=index+2, column=df_leads.columns.get_loc("Квалифицирован")+1).value = "Нет"
                wb.save("leads.xlsx")
                return False
    except Exception as e:
        print(f"⚠️ Не удалось прочитать leads.xlsx: {e}")

    # Если всё ок — ставим квалифицирован
    try:
        df_leads = pd.read_excel("leads.xlsx")
        match = df_leads[df_leads["ID"].astype(str) == user_id]
        if not match.empty:
            index = match.index[0]
            wb = openpyxl.load_workbook("leads.xlsx")
            ws = wb.active
            if "Квалифицирован" not in df_leads.columns:
                ws.cell(row=1, column=len(df_leads.columns)+1).value = "Квалифицирован"
            ws.cell(row=index+2, column=df_leads.columns.get_loc("Квалифицирован")+1).value = "Да"
            wb.save("leads.xlsx")
    except:
        pass

    return True

# === Очереди сообщений ===
message_queue = deque()
priority_queue = deque()
regular_queue = deque()
use_priority = True

last_sent_per_chat = {}
sent_messages_count = 0

# === Загрузка настроек ===
with open("settings.json", "r", encoding="utf-8") as f:
    config = json.load(f)

API_ID = config["API_ID"]
API_HASH = config["API_HASH"]
FORWARD_TO_USER = config["FORWARD_TO_USER"]
SESSION_NAME = "session_name"

client = TelegramClient(SESSION_NAME, API_ID, API_HASH)

# === Загрузка списков ===
def load_list(filename):
    with open(filename, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

TARGET_CHATS = list(map(int, load_list("target_chats.txt")))
GROUP_IDS = list(map(int, load_list("group_ids.txt")))
KEYWORDS_SET_1 = load_list("keywords_1.txt")
KEYWORDS_SET_2 = load_list("keywords_2.txt")
STOP_WORDS = load_list("stop_words.txt")

# === Загрузка базы данных ===
DB_FILE = "users_database.xlsx"
try:
    df = pd.read_excel(DB_FILE)
except FileNotFoundError:
    df = pd.DataFrame(columns=["Источник", "Ник", "Текст запроса", "Дата и время", "Был ли в базе"])

# === Обработка текста ===
def clean_text(text):
    return re.sub(r'[*_~`"]', "", text).lower()

def is_user_in_db(username):
    return username in df["Ник"].values

def can_write_user(username, raw_text):
    if username not in df["Ник"].values:
        recent_entries = df[df["Дата и время"] >= (datetime.now() - pd.Timedelta(days=5)).strftime("%Y-%m-%d %H:%M:%S")]
        if raw_text in recent_entries["Текст запроса"].values:
            return False
        return True

    user_entries = df[df["Ник"] == username]
    last_date_str = user_entries["Дата и время"].max()
    last_date = datetime.strptime(last_date_str, "%Y-%m-%d %H:%M:%S")
    if (datetime.now() - last_date).days < 30:
        return False

    recent_entries = df[df["Дата и время"] >= (datetime.now() - pd.Timedelta(days=5)).strftime("%Y-%m-%d %H:%M:%S")]
    if raw_text in recent_entries["Текст запроса"].values:
        return False

    return True

def log_message(source, username, text, was_in_db, user_id, sent=False):
    new_entry = {
        "Источник": source,
        "Ник": username,
        "Текст запроса": text,
        "Дата и время": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Был ли в базе": "Да" if was_in_db else "Нет",
        "Отправлено": "Да" if sent else "Нет"
    }
    global df
    new_entry["Ник"] = username if username else user_id
    df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
    with pd.ExcelWriter(DB_FILE, mode="w", engine="openpyxl") as writer:
        df.to_excel(writer, index=False)


def get_username_by_id(user_id: str) -> str:
    user_id = str(user_id)

    try:
        df_leads = pd.read_excel("leads.xlsx")
        match = df_leads[df_leads["ID"].astype(str) == user_id]
        if not match.empty and "Юзернейм" in match.columns:
            username = str(match.iloc[0]["Юзернейм"])
            print(f"🔎 Найден username для ID {user_id}: {username}")
            if isinstance(username, str) and username.strip():
                return username
    except Exception as e:
        print(f"⚠️ Не удалось прочитать leads.xlsx: {e}")

    try:
        df_users = pd.read_excel("users_database.xlsx", header=None)
        for i in range(len(df_users)):
            if str(df_users.iloc[i, 0]) == user_id:
                username = str(df_users.iloc[i, 1])
                if isinstance(username, str) and username.strip():
                    return username
    except Exception as e:
        print(f"⚠️ Не удалось прочитать users_database.xlsx: {e}")

    return ""

# === Отправка информации о лиде в свой чат ===
async def forward_lead_info(chat, username, raw_text, source="Чат", keywords=None):
    found_keywords = ", ".join(keywords) if keywords else "-"
    msg = (
        f"📍 Новый лид из {source}\n"
        f"🔗 Чат: {chat}\n"
        f"👤 @{username}\n"
        f"🔑 Ключевые слова: {found_keywords}\n\n"
        f"💬 <pre>{raw_text}</pre>"
    )
    await client.send_message(FORWARD_TO_USER, msg, parse_mode="html")

# === Загрузка лидов из leads.xlsx ===
def load_leads_from_excel():
    for _, row in df.iterrows():
        # 1. Пропускаем, если уже отправлено
        if str(row.get("Отправлено", "")).strip().lower() == "да":
            continue

        # 2. Пропускаем, если явно помечено как не квал
        if str(row.get("Квалификация", "")).strip().lower() == "не квал":
            continue
    try:
        df_leads = pd.read_excel("leads.xlsx")
        unsent = df_leads[
            (df_leads["Отправлено"] == "Нет") &
            (~df_leads.get("Квалификация", "").astype(str).str.lower().eq("не квал"))
        ]

        # Соберём уже добавленные лиды в очередь, чтобы не дублировать
        already_in_queue = {(uid, text) for uid, _, text in priority_queue}

        added_count = 0
        for _, row in unsent.iterrows():
            lead_key = (row["ID"], row["Текст"])
            if lead_key in already_in_queue:
                continue  # уже есть в очереди — пропускаем

            lead = (
                row["ID"],
                config["CHAT_INTRO_MESSAGE"].format(chat=row["Чат"]),
                row["Текст"]
            )
            priority_queue.append(lead)
            added_count += 1

        if added_count > 0:
            print(f"📥 Загружено {added_count} новых лидов из leads.xlsx")
    except Exception as e:
        print(f"⚠️ Ошибка при загрузке лидов из Excel: {e}")

async def send_intro_and_raw(client, recipient, intro_message, raw_text):
    try:
        await client.send_message(recipient, intro_message)
        print(f"✅ Вступительное сообщение отправлено -> {recipient}")
        await asyncio.sleep(1)
        await client.send_message(recipient, f"<pre>{raw_text}</pre>", parse_mode="html")
        print(f"✅ raw_text отправлен -> {recipient}")
        return True
    except Exception as e:
        print(f"❌ Ошибка при отправке intro/raw_text -> {recipient}: {e}")
        return False
    

# === Обработка очереди ===
async def process_queue():
    global sent_messages_count, use_priority
    send_success = False
    while True:
        if message_queue:
            user_id, intro_message, raw_text = message_queue.popleft()
        elif use_priority and priority_queue:
            user_id, intro_message, raw_text = priority_queue.popleft()
            use_priority = False
        elif regular_queue:
            user_id, intro_message, raw_text = regular_queue.popleft()
            use_priority = True
        elif priority_queue:
            user_id, intro_message, raw_text = priority_queue.popleft()
            use_priority = False
        else:
            await asyncio.sleep(10)
            continue

        try:
            await asyncio.sleep(random.randint(
                config["DELAY_BETWEEN_MESSAGES_MIN"],
                config["DELAY_BETWEEN_MESSAGES_MAX"]
            ))
            # Проверяем квалификацию
            if not is_qualified(user_id):
                print(f"❌ Пользователь {user_id} не квалифицирован, пропускаем")
                continue  # сразу переходим к следующему сообщению
            # ⛔️ Проверка на повтор текста — если уже отправляли такой текст другим пользователям
            try:
                df_leads = pd.read_excel("leads.xlsx")
                existing_texts = df_leads[df_leads["Отправлено"] == "Да"]["Текст"].tolist()
                if any(raw_text.strip() == t.strip() for t in existing_texts if len(t.strip()) > 50):
                    print(f"⚠️ Повтор текста — уже отправляли этот текст раньше, пропускаем")
                    # ⛔️ Помечаем, что лид не подходит (дубликат текста)
                    try:
                        idx = df_leads[(df_leads["ID"] == user_id) & (df_leads["Текст"] == raw_text)].index
                        if not idx.empty:
                            df_leads.loc[idx[0], "Отправлено"] = "Нет"
                            df_leads.loc[idx[0], "Квалификация"] = "Не квал"
                            df_leads.to_excel("leads.xlsx", index=False)
                            print(f"⛔️ Лид помечен как 'Не квал' — дубликат текста")
                    except Exception as e:
                        print(f"⚠️ Не удалось пометить лид как 'Не квал': {e}")
                         # ⚠️ Не удалось отправить — помечаем в leads.xlsx как "Не квал"
                        try:
                            df_leads = pd.read_excel("leads.xlsx")
                            idx = df_leads[
                                (df_leads["ID"].astype(str) == str(user_id)) &
                                (df_leads["Текст"] == raw_text)
                            ].index
                            if not idx.empty:
                                df_leads.loc[idx[0], "Квалификация"] = "Не квал"
                                df_leads.to_excel("leads.xlsx", index=False)
                                print(f"⛔️ Лид помечен как 'Не квал' — ошибка при отправке")
                        except Exception as save_err:
                            print(f"⚠️ Ошибка при записи 'Не квал': {save_err}")
                    continue
            except Exception as e:
                print(f"⚠️ Ошибка при проверке текста на дубликат: {e}")
            if config.get("SEND_MESSAGES", True):
                # ⛔️ Проверка: если уже отправляли этот лид — пропускаем
                try:
                    # 🔎 Проверка: уже отправляли такому ID или текст?
                    try:
                        df_leads = pd.read_excel("leads.xlsx")
                        df_users = pd.read_excel("users_database.xlsx", header=None) if os.path.exists("users_database.xlsx") else pd.DataFrame()

                        # Проверка ID
                        if str(user_id) in df_users[0].astype(str).values:
                            print(f"⛔️ Уже писали пользователю {user_id}, помечаем как 'Не квал'")
                            idx = df_leads[(df_leads["ID"].astype(str) == str(user_id)) & (df_leads["Текст"] == raw_text)].index
                            if not idx.empty:
                                df_leads.loc[idx[0], "Квалификация"] = "Не квал"
                                df_leads.to_excel("leads.xlsx", index=False)
                            continue

                        # Проверка текста
                        sent_texts = df_leads[df_leads["Отправлено"] == "Да"]["Текст"].dropna().astype(str).str.strip().tolist()
                        if any(raw_text.strip() == t.strip() for t in sent_texts if len(t.strip()) >= 60):
                            print("⛔️ Повтор текста — помечаем как 'Не квал'")
                            idx = df_leads[(df_leads["ID"].astype(str) == str(user_id)) & (df_leads["Текст"] == raw_text)].index
                            if not idx.empty:
                                df_leads.loc[idx[0], "Квалификация"] = "Не квал"
                                df_leads.to_excel("leads.xlsx", index=False)
                            continue

                        # Если всё ок — помечаем 'Квалификация' = 'Да'
                        idx = df_leads[(df_leads["ID"].astype(str) == str(user_id)) & (df_leads["Текст"] == raw_text)].index
                        if not idx.empty:
                            df_leads.loc[idx[0], "Квалификация"] = "Да"
                            df_leads.to_excel("leads.xlsx", index=False)

                    except Exception as e:
                        print(f"⚠️ Ошибка при проверке уникальности ID/текста: {e}")
                    df_leads = pd.read_excel("leads.xlsx")
                    mask = (df_leads["ID"] == user_id) & (df_leads["Текст"] == raw_text)
                    if not df_leads[mask].empty and df_leads.loc[mask, "Отправлено"].values[0] == "Да":
                        print(f"⚠️ Повтор — лид уже был отправлен (ID: {user_id}), пропускаем")
                        continue
                except Exception as e:
                    print(f"⚠️ Ошибка при проверке на дубликат в leads.xlsx: {e}")

                # ⛔️ Проверка в users_database.xlsx — если ID + текст уже логировались
                try:
                    df_users = pd.read_excel("users_database.xlsx", header=None)
                    if not df_users.empty:
                        id_col = df_users.columns[0]
                        text_col = df_users.columns[2]
                        duplicates = df_users[
                            (df_users[id_col].astype(str) == str(user_id)) &
                            (df_users[text_col] == raw_text)
                        ]
                        if not duplicates.empty:
                            print(f"⚠️ Повтор — пользователь {user_id} с этим текстом уже в users_database.xlsx, пропускаем")
                            continue
                except Exception as e:
                    print(f"⚠️ Ошибка при проверке на дубликат в users_database.xlsx: {e}")
                # Обновим users_database.xlsx (Квалифицирован = Да)
                try:
                    df_users = pd.read_excel("users_database.xlsx", header=None)
                    for i in range(len(df_users)):
                        if str(df_users.iloc[i, 0]) == str(user_id):
                            wb = openpyxl.load_workbook("users_database.xlsx")
                            ws = wb.active
                            col_count = ws.max_column
                            if col_count < 6:
                                ws.cell(row=i+1, column=6).value = "Да"
                            wb.save("users_database.xlsx")
                            break
                except Exception as e:
                    print(f"⚠️ Не удалось обновить квалификацию в users_database.xlsx: {e}")

                # Обновим leads.xlsx (если такой пользователь там есть)
                try:
                    df_leads = pd.read_excel("leads.xlsx")
                    match = df_leads[df_leads["ID"].astype(str) == str(user_id)]
                    if not match.empty:
                        index = match.index[0]
                        wb = openpyxl.load_workbook("leads.xlsx")
                        ws = wb.active
                        if "Квалифицирован" not in df_leads.columns:
                            ws.cell(row=1, column=len(df_leads.columns)+1).value = "Квалифицирован"
                        ws.cell(row=index+2, column=df_leads.columns.get_loc("Квалифицирован")+1).value = "Да"
                        wb.save("leads.xlsx")
                except Exception as e:
                    print(f"⚠️ Не удалось обновить квалификацию в leads.xlsx: {e}")
                

                try:
                    send_success = await send_intro_and_raw(client, user_id, intro_message, raw_text)
                    if not send_success:
                        raise Exception("❌ Не удалось отправить по user_id — пробуем username")
                except Exception as e:
                    print(f"⚠️ Не удалось отправить по ID: {e}")
                    username = get_username_by_id(user_id)
                    if username:
                        if not username.startswith("@"):
                            username = f"@{username}"
                        try:
                            send_success = await send_intro_and_raw(client, username, intro_message, raw_text)
                            print(f"✅ Успешно отправлено по username {username}")
                        except Exception as uname_err:
                            print(f"❌ Не удалось отправить по username: {uname_err}")

                    # ⛔️ ВНИМАНИЕ: только если send_success всё ещё False — ставим "не квал"
                    if not send_success:
                        try:
                            df_leads = pd.read_excel("leads.xlsx")
                            idx = df_leads[
                                (df_leads["ID"].astype(str) == str(user_id)) &
                                (df_leads["Текст"] == raw_text)
                            ].index
                            if not idx.empty:
                                df_leads.loc[idx[0], "Квалификация"] = "Не квал"
                                df_leads.to_excel("leads.xlsx", index=False)
                                print(f"⛔️ Лид помечен как 'Не квал' — ошибка при отправке")
                        except Exception as save_err:
                            print(f"⚠️ Ошибка при записи 'Не квал': {save_err}")


                if send_success:
                    print(f"[{datetime.now().strftime('%H:%M:%S')}] ✅ Отправлено пользователю {user_id}")
                else:
                    print(f"[{datetime.now().strftime('%H:%M:%S')}] ❌ Не удалось отправить пользователю {user_id}")
                
                # ✅ Лог во внутренний чат
                username = get_username_by_id(user_id) if 'get_username_by_id' in globals() else str(user_id)
                if username:
                    id_line = f"👤 ID: {user_id} @{username}"
                else:
                    id_line = f"👤 ID: {user_id}"

                await client.send_message(
                    FORWARD_TO_USER,
                    f"🟢 Лид ({'отправлен пользователю' if send_success else 'только лог'})\n\n"
                    f"{id_line}\n"
                    f"💬 <pre>{raw_text}</pre>\n"
                    f"📦 Из очереди (сообщение {'отправлено' if send_success else 'НЕ отправлено пользователю'})",
                    parse_mode="html"
                )

            else:
                # Попытка найти username в leads.xlsx или users_database.xlsx
                

                username = get_username_by_id(user_id)
                id_line = f"👤 @{username} (ID: {user_id})" if username else f"👤 ID: {user_id}"

                await client.send_message(
                    FORWARD_TO_USER,
                    f"🟢 Лид (только лог)\n\n"
                    f"{id_line}\n"
                    f"💬 <pre>{raw_text}</pre>\n"
                    f"📦 Из очереди (сообщение НЕ отправлено пользователю)",
                    parse_mode="html"
                )

                print(f"[{datetime.now().strftime('%H:%M:%S')}] ✏️ Лид логирован (без отправки)")
            if send_success:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] ✅ Отправлено пользователю {user_id}")
            else:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] ❌ Не удалось отправить пользователю {user_id}")
            sent_messages_count += 1

            try:
                df_leads = pd.read_excel("leads.xlsx")
                mask = (df_leads["ID"] == user_id) & (df_leads["Текст"] == raw_text)
                df_leads.loc[mask, "Отправлено"] = "Да"
                with pd.ExcelWriter("leads.xlsx", mode="w", engine="openpyxl") as writer:
                    df_leads.to_excel(writer, index=False)
            except Exception as e:
                print(f"⚠️ Не удалось обновить leads.xlsx: {e}")

        except FloodWaitError as e:
            print(f"⏳ FloodWaitError: подождём {e.seconds} секунд...")
            await asyncio.sleep(e.seconds)
        except PeerFloodError:
            print("🚨 PeerFloodError: Telegram может заблокировать. Бот остановлен.")
            exit()
        except Exception as e:
            print(f"⚠️ Ошибка при отправке пользователю {user_id}: {e}")
            if not send_success:
                print(f"⛔️ Лид не отправлен — помечен как 'Не квал'")
                # запись в leads.xlsx
                try:
                    df_leads = pd.read_excel("leads.xlsx")
                    idx = df_leads[
                        (df_leads["ID"].astype(str) == str(user_id)) &
                        (df_leads["Текст"] == raw_text)
                    ].index
                    if not idx.empty:
                        df_leads.loc[idx[0], "Квалификация"] = "Не квал"
                        df_leads.to_excel("leads.xlsx", index=False)
                        print("❌ Записали 'Не квал' в leads.xlsx")
                except Exception as e:
                    print(f"⚠️ Ошибка при записи 'Не квал': {e}")

# === Обработчик групп ===
@client.on(events.NewMessage(chats=GROUP_IDS))
async def group_handler(event):
    raw_text = event.message.message
    message_text = clean_text(raw_text)

    # Проверка на наличие стоп-слов
    if any(stop_word in message_text for stop_word in STOP_WORDS):
        return

    # Проверка на ключевые слова из обоих наборов
    if not (any(word in message_text for word in KEYWORDS_SET_1) and
            any(word in message_text for word in KEYWORDS_SET_2)):
        return

    # Найдём какие ключевые слова сработали
    matched_1 = [word for word in KEYWORDS_SET_1 if word in message_text]
    matched_2 = [word for word in KEYWORDS_SET_2 if word in message_text]
    matched_keywords = matched_1 + matched_2

    # Ищем все упомянутые usernames в сообщении
    usernames = set(re.findall(r"(?:@|https?://)?t\.me/([a-zA-Z0-9_]+)", raw_text)) | \
            set(re.findall(r"@([a-zA-Z0-9_]+)", raw_text))

    for uname in usernames:
        if is_user_in_db(uname):
            continue
        try:
            user = await client.get_entity(uname)
            chat_name = (await event.get_chat()).title
            intro = config["GROUP_INTRO_MESSAGE"].format(chat=chat_name)
            log_message("Группа", uname, raw_text, was_in_db=False, user_id=user.id, sent=False)
            message_queue.append((user.id, intro, raw_text))
            await forward_lead_info(chat_name, uname, raw_text, source="Чат", keywords=matched_keywords)
            print(f"📌 Найден в группе: @{uname}")
        except (UsernameInvalidError, UsernameNotOccupiedError, UserPrivacyRestrictedError) as e:
            print(f"⚠️ Не удалось получить @{uname}: {e}")
            continue

@client.on(events.NewMessage(chats=FORWARD_TO_USER))
async def command_handler(event):
    text = event.raw_text.strip().lower()

    # Если начинается с "бан ", добавим слово в стоп-лист
    if text.startswith("бан "):
        new_word = text.replace("бан ", "").strip()

        if not new_word:
            await event.reply("⚠️ Не указано слово для бана.")
            return

        # Проверим, не в стоп-листе ли уже
        if new_word in STOP_WORDS:
            await event.reply(f"⛔ Слово «{new_word}» уже в стоп-словах.")
            return

        # Добавим в файл
        with open("stop_words.txt", "a", encoding="utf-8") as f:
            f.write(f"\n{new_word}")
        STOP_WORDS.append(new_word)

        await event.reply(f"✅ Слово «{new_word}» добавлено в стоп-слова.")
        
    # Если начинается с "добавить слово ", добавим его в KEYWORDS_2
    elif text.startswith("добавить слово "):
        new_word = text.replace("добавить слово", "").strip()

        if not new_word:
            await event.reply("⚠️ Не указано слово для добавления в KEYWORDS_2.")
            return

        if new_word in KEYWORDS_SET_2:
            await event.reply(f"⛔ Слово «{new_word}» уже есть в KEYWORDS_2.")
            return

        with open("keywords_2.txt", "a", encoding="utf-8") as f:
            f.write(f"\n{new_word}")
        KEYWORDS_SET_2.append(new_word)

        await event.reply(f"✅ Слово «{new_word}» добавлено в KEYWORDS_2.")


@client.on(events.NewMessage(chats=TARGET_CHATS))
async def chat_handler(event):
    chat_name = (await event.get_chat()).title
    now = datetime.now()

    last_time = last_sent_per_chat.get(chat_name)
    if last_time and (now - last_time).total_seconds() < 900:  # 15 минут
        return  # Пропускаем, если слишком рано

    # Если не вернулись — обновляем таймер
    last_sent_per_chat[chat_name] = now
    raw_text = event.message.message
    message_text = clean_text(raw_text)

    # Проверка на наличие стоп-слов
    if any(stop_word in message_text for stop_word in STOP_WORDS):
        return

    # Проверка на ключевые слова из обоих наборов
    if not (any(word in message_text for word in KEYWORDS_SET_1) and
            any(word in message_text for word in KEYWORDS_SET_2)):
        return

    # Найдём какие ключевые слова сработали
    matched_1 = [word for word in KEYWORDS_SET_1 if word in message_text]
    matched_2 = [word for word in KEYWORDS_SET_2 if word in message_text]
    matched_keywords = matched_1 + matched_2

    # Ищем usernames в сообщении
    usernames = set(re.findall(r"(?:@|https?://)?t\.me/([a-zA-Z0-9_]+)", raw_text)) | \
            set(re.findall(r"@([a-zA-Z0-9_]+)", raw_text))

    # Если есть usernames — пишем им
    if usernames:
        for uname in usernames:
            if is_user_in_db(uname):
                continue
            try:
                user = await client.get_entity(uname)
                chat_name = (await event.get_chat()).title
                intro = config["CHAT_INTRO_MESSAGE"].format(chat=chat_name)
                log_message(chat_name, uname, raw_text, was_in_db=False, user_id=user.id, sent=False)
                message_queue.append((user.id, intro, raw_text))
                await forward_lead_info(chat_name, uname, raw_text, source="Чат", keywords=matched_keywords)
                print(f"🟢 Лид из чата (упоминание): @{uname} — добавлен в очередь")
            except (UsernameInvalidError, UsernameNotOccupiedError, UserPrivacyRestrictedError) as e:
                print(f"⚠️ Не удалось получить @{uname}: {e}")
                continue
    else:
        # Если usernames нет — пишем отправителю
        sender = await event.get_sender()
        if sender is None:
            return
        username = sender.username or sender.id
        if not can_write_user(username, raw_text):
            return
        chat_name = (await event.get_chat()).title
        intro = config["CHAT_INTRO_MESSAGE"].format(chat=chat_name)
        log_message(chat_name, username, raw_text, was_in_db=False, user_id=sender.id, sent=False)
        # Обновим users_database.xlsx (Квалифицирован = Да)
        message_queue.append((sender.id, intro, raw_text))
        print(f"🟢 Лид из чата (отправитель): @{username} — добавлен в очередь")




async def periodic_leads_loader():
    while True:
        print("🔁 Проверка новых лидов в leads.xlsx...")
        load_leads_from_excel()
        await asyncio.sleep(300)  # каждые 5 минут
async def periodic_stop_words_updater():
    global STOP_WORDS
    while True:
        try:
            with open("stop_words.txt", "r", encoding="utf-8") as f:
                STOP_WORDS = [line.strip() for line in f if line.strip()]
            print("🔄 Стоп-слова обновлены из файла.")
        except Exception as e:
            print(f"⚠️ Не удалось обновить стоп-слова: {e}")
        await asyncio.sleep(300)  # каждые 5 минут

def update_qualification_for_all():
    try:
        df_leads = pd.read_excel("leads.xlsx")
        if "Квалифицирован" not in df_leads.columns:
            df_leads["Квалифицирован"] = ""

        for index, row in df_leads.iterrows():
            отправлено = str(row.get("Отправлено", "")).strip().lower()
            квалификация = str(row.get("Квалификация", "")).strip().lower()

            if отправлено == "да" or квалификация == "не квал":
                df_leads.at[index, "Квалифицирован"] = "Нет"
            else:
                df_leads.at[index, "Квалифицирован"] = "Да"

        df_leads.to_excel("leads.xlsx", index=False)
        print("✅ Квалификация лидов обновлена в leads.xlsx")

    except Exception as e:
        print(f"⚠️ Ошибка при анализе квалификации лидов: {e}")


# === Запуск ===
async def main():
    await client.start()
    print("🤖 Бот запущен и мониторит чаты, группы и лиды из Excel...")
    update_qualification_for_all()
    load_leads_from_excel()
    asyncio.create_task(process_queue())
    asyncio.create_task(periodic_leads_loader())
    asyncio.create_task(periodic_stop_words_updater())
    try:
        await client.run_until_disconnected()
    except Exception as e:
        if "PersistentTimestampOutdatedError" not in str(e):
            print(f"🚨 Ошибка: {e}")
        else:
            pass  # молча игнорируем PersistentTimestampOutdatedError


    

asyncio.run(main())