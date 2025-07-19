# 1. Импорты

# === Стандартная библиотека ===
import asyncio
import json
import logging
import os
import re
import openpyxl
from datetime import datetime, timedelta
from typing import List

# === Сторонние библиотеки ===
import pandas as pd
import matplotlib.pyplot as plt
from telethon import TelegramClient, events

# === Локальные модули ===
# Пока не используются

# 2. Константы и глобальные переменные


# === Пути к файлам и директориям ===
PROJECT_PATH = "projects"
STOP_WORDS_PATH = "core/stop_words.txt"
SENT_MESSAGES_PATH = "core/sent_messages.xlsx"
KEYWORDS1_MESSAGES_PATH = "core/keywords1_messages.xlsx"
STATS_HOURLY_MAIN_PATH = "core/stats_hourly_main.json"
STATS_HOURLY_TEST_PATH = "core/stats_hourly_test.json"
TEST_STATUS_PATH = "core/testprojectstatus.json"
DEFAULT_TEST_LIMIT = 10

# === Ограничения и интервалы ===
MAX_MESSAGE_LENGTH = 500
STOPWORDS_REFRESH_INTERVAL = 3600  # сек (1 час)
CHATS_REFRESH_INTERVAL = 3600      # сек (1 час)

# === Глобальные таймеры для автообновлений ===
last_stopwords_update = datetime.min  # Последнее обновление стоп-слов
last_sources_update = datetime.min    # Последнее обновление чатов

# === Служебные глобальные переменные ===
sent_message_ids = set()                 # ID уже отправленных сообщений
sent_texts_keywords1 = set()            # Тексты сообщений по KEYWORDS_1 (для исключения дублей)

# === Счётчики для отчёта (сброс каждый час) ===
messages_analyzed = 0                   # Всего сообщений обработано
messages_matched = 0                    # Подошло по KEYWORDS_1 + KEYWORDS_2
messages_matched_keywords1 = 0          # Подошло по только KEYWORDS_1

# === Загрузка стоп-слов ===
try:
    STOP_WORDS = [line.strip().lower() for line in open(STOP_WORDS_PATH, encoding="utf-8") if line.strip()]
    logging.info(f"✅ Загружено {len(STOP_WORDS)} стоп-слов")
except Exception as e:
    STOP_WORDS = []
    logging.error(f"❌ Не удалось загрузить стоп-слова: {e}")

# === Загрузка таблицы с уже отправленными текстами ===
try:
    df_sent = pd.read_excel(SENT_MESSAGES_PATH)
    sent_texts_keywords1 = set(df_sent["Текст"].dropna().str.lower())
    logging.info(f"✅ Загружено {len(sent_texts_keywords1)} отправленных текстов из sent_messages.xlsx")
except FileNotFoundError:
    df_sent = pd.DataFrame(columns=["Дата", "Текст"])
    logging.warning("⚠️ Файл sent_messages.xlsx не найден. Создана новая таблица.")




# 3. Загрузка конфигов и инициализация клиента

# === Загрузка settings.json ===
try:
    with open("core/settings.json", "r", encoding="utf-8") as f:
        config = json.load(f)
        logging.info("✅ Конфигурация успешно загружена из settings.json")
except Exception as e:
    logging.error(f"❌ Ошибка при загрузке settings.json: {e}")
    config = {}

# === Чтение ключевых параметров из config ===
API_ID = config.get("API_ID")
API_HASH = config.get("API_HASH")
BOT_TOKEN = config.get("BOT_TOKEN")

# Попытка определить активный проект
project_config_path = "core/project_config.json"
if os.path.exists(project_config_path):
    try:
        with open(project_config_path, "r", encoding="utf-8") as f:
            project_cfg = json.load(f)
        PROJECT = project_cfg.get("project", config.get("PROJECT", "default"))
    except Exception as e:
        logging.error(f"❌ Ошибка при чтении project_config.json: {e}")
        PROJECT = config.get("PROJECT", "default")
else:
    PROJECT = config.get("PROJECT", "default")

TEST_LIMIT = config.get("test_limit", DEFAULT_TEST_LIMIT)
TEST_MODE = config.get("test_mode", False)

# === Пути до файлов проекта ===
KEYWORDS_1_PATH = f"{PROJECT_PATH}/{PROJECT}/keywords_1.txt"
KEYWORDS_2_PATH = f"{PROJECT_PATH}/{PROJECT}/keywords_2.txt"
STOP_WORDS_PROJECT_PATH = f"{PROJECT_PATH}/{PROJECT}/stop_words.txt"

def load_words_from_file(path: str) -> List[str]:
    """Загружает список слов из текстового файла"""
    try:
        with open(path, encoding="utf-8") as f:
            words = [line.strip().lower() for line in f if line.strip()]
        logging.info(f"✅ Загружено {len(words)} слов из {path}")
        return words
    except FileNotFoundError:
        logging.warning(f"⚠️ Файл {path} не найден")
        return []
    except Exception as e:
        logging.error(f"❌ Ошибка при чтении файла {path}: {e}")
        return []

# === Переменные проекта ===
KEYWORDS_1 = load_words_from_file(KEYWORDS_1_PATH)
KEYWORDS_2 = load_words_from_file(KEYWORDS_2_PATH)
STOP_WORDS_PROJECT = load_words_from_file(STOP_WORDS_PROJECT_PATH)

# === ID админов для отчётов ===
ADMIN_CHAT_ID = config.get("ADMIN_CHAT_ID")
ADMIN_TEST_CHAT_ID = config.get("ADMIN_TEST_CHAT_ID") 

# === Проверка критически важных значений ===
if not all([API_ID, API_HASH]):
    logging.error("❌ В settings.json отсутствуют необходимые параметры!")
else:
    logging.info(f"🔧 Настройки проекта '{PROJECT}' успешно загружены (test_mode={TEST_MODE})")

# === Пути до файлов проекта ===
KEYWORDS_1_PATH = f"{PROJECT_PATH}/{PROJECT}/keywords_1.txt"
KEYWORDS_2_PATH = f"{PROJECT_PATH}/{PROJECT}/keywords_2.txt"
STOP_WORDS_PROJECT_PATH = f"{PROJECT_PATH}/{PROJECT}/stop_words.txt"
TARGET_CHATS_MAIN_PATH = f"{PROJECT_PATH}/{PROJECT}/target_chats.txt"
TARGET_CHATS_TEST_PATH = f"{PROJECT_PATH}/{PROJECT}/target_chats_test.txt"

# === Инициализация Telegram клиента ===
SESSION_NAME = config.get("SESSION_NAME", "session_name")
client = TelegramClient(SESSION_NAME, API_ID, API_HASH)

# === Лог успешной инициализации ===
logging.info("📲 TelegramClient успешно инициализирован")

# === ID ботов ===
def load_bot_ids(path: str = "core/bot_ids.txt") -> list[int]:
    """Загружает список Telegram bot ID из файла"""
    try:
        with open(path, encoding="utf-8") as f:
            ids = [int(line.strip()) for line in f if line.strip().isdigit()]
        logging.info(f"✅ Загружено {len(ids)} bot ID из {path}")
        return ids
    except FileNotFoundError:
        logging.warning(f"⚠️ Файл {path} не найден")
        return []
    except Exception as e:
        logging.error(f"❌ Ошибка при загрузке bot ID: {e}")
        return []

TELEGRAM_BOT_IDS = load_bot_ids()
















# 4. Вспомогательные функции

def load_words_from_file(path: str) -> list[str]:
    """Загружает список слов из текстового файла"""
    try:
        with open(path, encoding="utf-8") as f:
            words = [line.strip().lower() for line in f if line.strip()]
        logging.info(f"✅ Загружено {len(words)} слов из {path}")
        return words
    except FileNotFoundError:
        logging.warning(f"⚠️ Файл {path} не найден")
        return []
    except Exception as e:
        logging.error(f"❌ Ошибка при чтении файла {path}: {e}")
        return []

def normalize_text(text: str) -> str:
    """Очищает текст сообщения от ссылок, символов и приводит к нижнему регистру"""
    clean = re.sub(r"[^\w\s@.]", " ", text)
    return re.sub(r"\s+", " ", clean).strip().lower()

def update_hourly_statistics(project: str, test_mode: bool):
    """Обновляет почасовую статистику для отчёта"""
    stats_file = STATS_HOURLY_TEST_PATH if test_mode else STATS_HOURLY_MAIN_PATH
    now = datetime.now().replace(minute=0, second=0, microsecond=0)
    try:
        if os.path.exists(stats_file):
            with open(stats_file, "r", encoding="utf-8") as f:
                stats = json.load(f)
        else:
            stats = {}

        now_str = now.isoformat()
        stats[now_str] = stats.get(now_str, 0) + 1

        with open(stats_file, "w", encoding="utf-8") as f:
            json.dump(stats, f, ensure_ascii=False, indent=2)
        logging.info(f"📊 Обновлена статистика: {now_str} (+1)")
    except Exception as e:
        logging.error(f"❌ Ошибка при обновлении статистики: {e}")

def is_duplicate_message(text: str) -> bool:
    """Проверяет, было ли это сообщение уже отправлено (по тексту)"""
    return text in sent_texts_keywords1

def load_json_file(file_path: str) -> dict:
    """Загружает JSON-файл, возвращает пустой словарь при ошибке"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"❌ Ошибка при загрузке JSON из {file_path}: {e}")
        return {}

def save_message_to_excel(message: str, file_path: str):
    """Сохраняет сообщение в Excel-файл (добавляет в конец)"""
    try:
        df = pd.read_excel(file_path) if os.path.exists(file_path) else pd.DataFrame(columns=["Дата", "Текст"])
        new_row = {"Дата": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Текст": message}
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        df.to_excel(file_path, index=False)
        logging.info(f"💾 Сообщение сохранено в {file_path}")
    except Exception as e:
        logging.error(f"❌ Ошибка при сохранении сообщения: {e}")

def read_and_clear_hourly_stats(file_path: str) -> dict:
    """Читает и очищает json-файл почасовой статистики"""
    try:
        if os.path.exists(file_path):
            with open(file_path, "r", encoding="utf-8") as f:
                stats = json.load(f)
            os.remove(file_path)
            logging.info(f"✅ Прочитана и очищена статистика из {file_path}")
            return stats
        else:
            return {}
    except Exception as e:
        logging.error(f"❌ Ошибка при чтении hourly stats: {e}")
        return {}

async def log_lead_to_admin_chat(event, matched_keywords: list[str]):
    """Логирует найденного лида в админский чат"""
    try:
        text = event.message.message or ""
        user = await event.get_sender()
        if user.id in TELEGRAM_BOT_IDS:
            username_from_button = extract_username_from_button(event)
            if username_from_button:
                raw_text = f"{username_from_button}\n{raw_text}"
            raw_text = cut_text_before_symbol(raw_text)
        username = f"@{user.username}" if user.username else "Без username"
        user_id = user.id

        message = (
            f"🧪 Найден лид\n"
            f"👤 {username} (ID: {user_id})\n"
            f"📦 {', '.join(matched_keywords)}\n\n"
            f"{text}"
        )

        chat_id = ADMIN_TEST_CHAT_ID if TEST_MODE else ADMIN_CHAT_ID
        await client.send_message(chat_id, message)

    except Exception as e:
        logging.error(f"❌ Ошибка при логировании лида: {e}")

def extract_username_from_button(event) -> str | None:
    """
    Извлекает username из кнопки сообщения (если она начинается с ✍️ и содержит ссылку)
    """
    try:
        buttons = event.message.reply_markup.rows
        for row in buttons:
            for button in row.buttons:
                if button.text.startswith("✍️") and "t.me/" in button.url:
                    # Вытащить username из ссылки
                    return "@" + button.url.split("t.me/")[1].split("?")[0]
    except Exception as e:
        logging.warning(f"❌ Ошибка при разборе кнопки: {e}")
    return None

def cut_text_before_symbol(text: str, symbol: str = "➖") -> str:
    """Обрезает текст по указанному символу (по умолчанию — '➖')"""
    return text.split(symbol)[0].strip()

# сохранение лидов в ексель проектов
def save_lead_to_project_excel(project_path, sender_id, username, matched_keywords, text, source):
    file_path = os.path.join(project_path, "users_database.xlsx")
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Дата", "ID", "Username", "Ключевые слова", "Сообщение", "Источник"])
    else:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

    # Проверка на дублирование
    existing_texts = [row[4].value for row in worksheet.iter_rows(min_row=2) if row[4].value]
    if text in existing_texts:
        return

    worksheet.append([timestamp, sender_id, username, ", ".join(matched_keywords), text, source])
    workbook.save(file_path)





















# 5. Обработчики событий

# === Обработчик входящих сообщений из чатов ===
@client.on(events.NewMessage(incoming=True))
async def handle_incoming_message(event):
    """Обрабатывает входящее сообщение, применяет фильтры и сохраняет лид"""
    global messages_analyzed, messages_matched, messages_matched_keywords1

    try:
        # Пропускаем ЛС и каналы
        if not event.is_group:
            return

        chat_id = event.chat_id
        
        # Проверка: чат должен быть в целевых
        target_chats_path = TARGET_CHATS_TEST_PATH if TEST_MODE else TARGET_CHATS_MAIN_PATH
        try:
            with open(target_chats_path, encoding="utf-8") as f:
                allowed_chats = set(int(line.strip()) for line in f if line.strip().isdigit())
        except Exception as e:
            logging.warning(f"⚠️ Не удалось загрузить список чатов: {e}")
            allowed_chats = set()

        if chat_id not in allowed_chats:
            logging.debug(f"⏭️ Чат {chat_id} не в целевом списке — пропуск")
            return
        raw_text = event.message.message or ""
        if not raw_text.strip() or len(raw_text) > MAX_MESSAGE_LENGTH:
            logging.debug(f"⏭️ Пропущено: пустое или слишком длинное сообщение ({len(raw_text)} символов)")
            return

        messages_analyzed += 1
        normalized = normalize_text(raw_text)

        # Проверка стоп-слов
        if any(word in normalized for word in STOP_WORDS + STOP_WORDS_PROJECT):
            logging.info(f"🚫 Стоп-слово в сообщении: {raw_text[:60]}...")
            return

        # Проверка на KEYWORDS_1
        if not any(kw in normalized for kw in KEYWORDS_1):
            return

        messages_matched_keywords1 += 1

        # Проверка на KEYWORDS_2
        matched_keywords2 = [kw for kw in KEYWORDS_2 if kw in normalized]
        if not matched_keywords2:
            return

        messages_matched += 1

        # Проверка на дубли
        if is_duplicate_message(normalized):
            logging.info("🔁 Дубликат сообщения — уже отправлялось")
            return

        # Сохранение как лид
        save_message_to_excel(raw_text, SENT_MESSAGES_PATH)
        sent_texts_keywords1.add(normalized)

        update_hourly_statistics(PROJECT, TEST_MODE)

        logging.info(f"✅ Найден лид: {raw_text[:60]}...")
        logging.info(f"📦 Ключевые слова: {matched_keywords2}")

        # Если не test_mode — логика отправки
        if not TEST_MODE:
            await log_lead_to_admin_chat(event, matched_keywords2)
        else:
            # Тестовый режим: проверка и логика
            test_state = increment_test_leads(PROJECT)
            if test_state["limit_reached"]:
                logging.warning("⚠️ Достигнут лимит тестовых лидов — лид не отправляется")
                return

            try:
                await client.forward_messages(ADMIN_TEST_CHAT_ID, event.message)
                logging.info("📤 Сообщение переслано в тестовый чат")
                if test_state["sent"] == TEST_LIMIT:
                    await client.send_message(ADMIN_TEST_CHAT_ID, f"📛 Проект '{PROJECT}' достиг тестового лимита в {TEST_LIMIT} лидов")
            except Exception as e:
                logging.error(f"❌ Ошибка при пересылке тестового лида: {e}")

    except Exception as e:
        logging.error(f"❌ Ошибка в обработчике handle_incoming_message: {e}")

# тестовый режим

def load_test_status():
    if os.path.exists(TEST_STATUS_PATH):
        with open(TEST_STATUS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    else:
        return {}

def save_test_status(status):
    with open(TEST_STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=2)

def increment_test_leads(project_name):
    status = load_test_status()
    status.setdefault(project_name, {"sent": 0, "limit_reached": False})

    if not status[project_name]["limit_reached"]:
        status[project_name]["sent"] += 1
        if status[project_name]["sent"] >= TEST_LIMIT:
            status[project_name]["limit_reached"] = True
    save_test_status(status)
    return status[project_name]

























# 6. Планировщики (hourly report, график)

async def send_hourly_report():
    """Формирует и отправляет почасовой отчёт по активности"""
    stats_file = STATS_HOURLY_TEST_PATH if TEST_MODE else STATS_HOURLY_MAIN_PATH
    stats = read_and_clear_hourly_stats(stats_file)

    if not stats:
        logging.info("📭 Нет данных для почасового отчета — пропуск")
        return

    try:
        report_lines = ["📊 Почасовой отчет:"]
        total = 0
        for hour, count in sorted(stats.items()):
            dt = datetime.fromisoformat(hour).strftime("%H:%M")
            report_lines.append(f"• {dt} — {count} сообщений")
            total += count

        report_lines.append(f"\nВсего: {total} сообщений")

        report_text = "\n".join(report_lines)
        chat_id = ADMIN_TEST_CHAT_ID if TEST_MODE else ADMIN_CHAT_ID

        await client.send_message(chat_id, report_text)
        logging.info("✅ Почасовой отчёт отправлен")

    except Exception as e:
        logging.error(f"❌ Ошибка при отправке почасового отчёта: {e}")


def build_activity_plot(stats: dict[str, int], output_path: str):
    """Строит и сохраняет график активности по часам"""
    try:
        if not stats:
            logging.info("📉 Нет данных для построения графика")
            return

        hours = [datetime.fromisoformat(h).strftime("%H:%M") for h in sorted(stats)]
        values = [stats[h] for h in sorted(stats)]

        plt.figure(figsize=(10, 5))
        plt.plot(hours, values, marker="o")
        plt.xticks(rotation=45)
        plt.title("График активности по часам")
        plt.tight_layout()
        plt.savefig(output_path)
        plt.close()

        logging.info(f"🖼️ График активности сохранён в {output_path}")
    except Exception as e:
        logging.error(f"❌ Ошибка при построении графика: {e}")


async def send_activity_plot():
    """Создаёт график активности и отправляет его в чат"""
    stats_file = STATS_HOURLY_TEST_PATH if TEST_MODE else STATS_HOURLY_MAIN_PATH
    image_path = f"{PROJECT_PATH}/{PROJECT}/activity_plot.png"

    stats = read_and_clear_hourly_stats(stats_file)
    if not stats:
        logging.info("📭 Нет данных для графика активности — пропуск")
        return

    build_activity_plot(stats, image_path)

    try:
        chat_id = ADMIN_TEST_CHAT_ID if TEST_MODE else ADMIN_CHAT_ID
        await client.send_file(chat_id, image_path, caption="📈 График активности за день")
        logging.info("✅ График активности отправлен")
    except Exception as e:
        logging.error(f"❌ Ошибка при отправке графика: {e}")


async def scheduler_loop():
    """Асинхронный цикл планировщика: запускает задачи по времени"""
    while True:
        now = datetime.now()

        # Каждый час — отчёт
        if now.minute == 0:
            await send_hourly_report()

        # Каждый день в 22:00 — график
        if now.hour == 22 and now.minute == 0:
            await send_activity_plot()

        if now.hour == 18 and now.minute == 0:
            await send_daily_leads_report()

        if now.weekday() == 4 and now.hour == 19 and now.minute == 0:
            await send_weekly_leads_report()

        await asyncio.sleep(60)

async def send_daily_leads_report():
    """Отправляет дневной отчёт, если есть лиды > 0"""
    stats_file = STATS_HOURLY_TEST_PATH if TEST_MODE else STATS_HOURLY_MAIN_PATH
    try:
        if not os.path.exists(stats_file):
            return
        with open(stats_file, "r", encoding="utf-8") as f:
            stats = json.load(f)

        # Считаем за последние 24 часа (от 18:00 до 18:00)
        cutoff = datetime.now() - timedelta(days=1)
        total = sum(v for k, v in stats.items() if datetime.fromisoformat(k) >= cutoff)

        if total < 1:
            logging.info("📭 Лидов за сутки < 1 — дневной отчет не отправляется")
            return

        # Загрузка project_config.json для имени проекта и ID чата
        project_config_path = f"{PROJECT_PATH}/{PROJECT}/project_config.json"
        if not os.path.exists(project_config_path):
            logging.warning("⚠️ project_config.json не найден")
            return

        with open(project_config_path, "r", encoding="utf-8") as f:
            project_config = json.load(f)

        chat_id = project_config.get("report_chat")
        project_name = project_config.get("project_name", PROJECT)

        if chat_id:
            text = f"📊 Добрый вечер. За сегодня для вашего проекта найдено {total} лидов"
            await client.send_message(chat_id, text)
            logging.info(f"📤 Отправлен дневной отчет ({total} лидов)")
    except Exception as e:
        logging.error(f"❌ Ошибка в send_daily_leads_report: {e}")


async def send_weekly_leads_report():
    """Отправляет недельный отчёт по пятницам"""
    stats_file = STATS_HOURLY_TEST_PATH if TEST_MODE else STATS_HOURLY_MAIN_PATH
    try:
        if not os.path.exists(stats_file):
            return
        with open(stats_file, "r", encoding="utf-8") as f:
            stats = json.load(f)

        # Считаем за последние 7 дней
        cutoff = datetime.now() - timedelta(days=7)
        total = sum(v for k, v in stats.items() if datetime.fromisoformat(k) >= cutoff)

        if total < 1:
            logging.info("📭 Лидов за неделю < 1 — недельный отчет не отправляется")
            return

        project_config_path = f"{PROJECT_PATH}/{PROJECT}/project_config.json"
        if not os.path.exists(project_config_path):
            logging.warning("⚠️ project_config.json не найден")
            return

        with open(project_config_path, "r", encoding="utf-8") as f:
            project_config = json.load(f)

        chat_id = project_config.get("report_chat")
        project_name = project_config.get("project_name", PROJECT)

        if chat_id:
            text = f"📈 За неделю в проекте '{project_name}' найдено {total} лидов"
            await client.send_message(chat_id, text)
            logging.info(f"📤 Отправлен недельный отчет ({total} лидов)")
    except Exception as e:
        logging.error(f"❌ Ошибка в send_weekly_leads_report: {e}")














# 7. Основная функция запуска

def setup_logging():
    """Настраивает логирование в консоль и файл"""
    log_formatter = logging.Formatter("%(asctime)s — %(levelname)s — %(message)s")
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Лог в файл
    file_handler = logging.FileHandler("search_bot.log", encoding="utf-8")
    file_handler.setFormatter(log_formatter)
    logger.addHandler(file_handler)

    # Лог в консоль
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(log_formatter)
    logger.addHandler(console_handler)

    logging.info("🔧 Логирование настроено")


def initialize_globals():
    """Инициализирует глобальные переменные, необходимые перед запуском"""
    global sent_texts_keywords1
    sent_texts_keywords1 = set()

    logging.info("🔁 Глобальные переменные инициализированы")


async def main():
    """Главная асинхронная функция запуска проекта"""
    setup_logging()
    initialize_globals()

    logging.info("🚀 Запуск Telegram клиента...")

    try:
        await client.start()
        logging.info("✅ Telegram клиент успешно запущен")

        # Запуск планировщика
        asyncio.create_task(scheduler_loop())
        logging.info("⏰ Планировщик задач запущен")

        # Ожидание событий
        logging.info(f"🟢 Бот запущен в режиме {'ТЕСТ' if TEST_MODE else 'БОЕВОМ'}")
        await client.run_until_disconnected()

    except Exception as e:
        logging.critical(f"❌ Критическая ошибка при запуске: {e}")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logging.info("🛑 Завершение работы по Ctrl+C")
